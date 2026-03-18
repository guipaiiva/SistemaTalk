"""
PARCIAL DO DIA — Grupo Talk (Linx Microvix)

Fluxo:
  1. Busca vendas do dia atual de todas as lojas via API Microvix
  2. Gera relatório em Excel (PARCIAL_DIA.xlsx)
  3. Converte para imagem PNG (PARCIAL_DIA.png)
  4. Envia via WhatsApp (Evolution API)

Regras:
  - Usa a mesma lógica de faturamento da Tendência v8.1
  - Lojas sem nenhuma venda no dia aparecem como "COLADA"
  - Sem meta, sem tendência — apenas nome da loja e valor do dia
"""

import os
import sys
import requests
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from PIL import Image, ImageDraw, ImageFont

# ============================================================
# CARREGA VARIÁVEIS DO ARQUIVO .env
# ============================================================
# Sobe dois níveis a partir deste arquivo para achar o .env na raiz do projeto
_ROOT = Path(__file__).resolve().parent.parent.parent
load_dotenv(_ROOT / ".env")

URL    = os.getenv("MICROVIX_URL",   "https://webapi.microvix.com.br/1.0/api/integracao")
USUARIO = os.getenv("MICROVIX_USER")
SENHA   = os.getenv("MICROVIX_SENHA")
CHAVE   = os.getenv("MICROVIX_CHAVE")

EVOLUTION_URL      = os.getenv("EVOLUTION_URL")
EVOLUTION_APIKEY   = os.getenv("EVOLUTION_KEY")
EVOLUTION_INSTANCE = os.getenv("EVOLUTION_INSTANCIA")

# Destinatários separados por vírgula no .env
# Ex: WHATSAPP_DESTINATARIOS=5561999734458,120363340293880028@g.us
_dest_raw = os.getenv("WHATSAPP_DESTINATARIOS", "")
WHATSAPP_DESTINATARIOS = [d.strip() for d in _dest_raw.split(",") if d.strip()]

# ============================================================
# PASTA DE SAÍDA (backend/output/)
# ============================================================
OUTPUT_DIR = Path(__file__).resolve().parent.parent / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

OUTPUT_EXCEL = str(OUTPUT_DIR / "PARCIAL_DIA.xlsx")
OUTPUT_IMAGE = str(OUTPUT_DIR / "PARCIAL_DIA.png")

# ============================================================
# MAPEAMENTO DE LOJAS
# ============================================================

# CNPJ → nome da loja (informação de negócio, não é segredo)
LOJAS_ATIVAS = {
    "04433822000117": "TALK BOULEVARD",
    "18410451000162": "LIG PKS",
    "18410451000243": "TALK PATIO",
    "41401165000233": "TALK DESEMBARQUE",
    "41401165000152": "TALK AERO",
    "04433822000389": "TALK BRASILIA",
    "05143548000376": "TALK TERRACO",
    "39281262000107": "TECH FLORIPA",
    "43785785000168": "TALK TGS",
    "45860610000101": "TALK PKS QUIOSQUE",
    "41401165000314": "TALK CONECTOR",
    "41401165000403": "TALK EMBARQUE",
    "43785785000249": "TECH PKS",
    "45860610000284": "TALK CNB",
    "41401165000586": "JK",
    "41401165000667": "MANHATTAN",
}

# Ordem de exibição por rota
ROTA_NORTE = [
    "TALK BOULEVARD",
    "TALK PATIO",
    "TALK BRASILIA",
    "TECH FLORIPA",
    "TALK CNB",
    "JK",
]
ROTA_SUL = [
    "LIG PKS",
    "TALK DESEMBARQUE",
    "TALK AERO",
    "TALK TERRACO",
    "TALK TGS",
    "TALK PKS QUIOSQUE",
    "TALK CONECTOR",
    "TALK EMBARQUE",
    "TECH PKS",
    "MANHATTAN",
]

# ============================================================
# CONFIGURAÇÕES VISUAIS DA IMAGEM
# ============================================================
COL_WIDTHS = [260, 200]
ROW_HEIGHT = 36
PAD_X      = 24
PAD_Y      = 24

COR_HEADER       = (47, 79, 127)
COR_COLADA       = (255, 100, 100)
COR_VALOR        = (235, 245, 255)
COR_ROTA         = (180, 200, 230)
COR_CINZA        = (221, 221, 221)
COR_BORDA        = (160, 160, 160)
COR_BRANCO       = (255, 255, 255)
COR_PRETO        = (30, 30, 30)
COR_TEXTO_HEADER = (255, 255, 255)
COR_TEXTO_COLADA = (255, 255, 255)


# ============================================================
# ETAPA 1 — BUSCA NA API MICROVIX
# ============================================================

# Template XML da requisição Microvix
XML_TEMPLATE = """<?xml version="1.0" encoding="utf-8"?>
<LinxMicrovix>
  <Authentication user="{usuario}" password="{senha}"/>
  <ResponseFormat>xml</ResponseFormat>
  <Command>
    <Name>LinxMovimento</Name>
    <Parameters>
      <Parameter id="chave">{chave}</Parameter>
      <Parameter id="cnpjEmp">{cnpj}</Parameter>
      <Parameter id="data_inicial">{data_ini}</Parameter>
      <Parameter id="data_fim">{data_fim}</Parameter>
      <Parameter id="timestamp">0</Parameter>
    </Parameters>
  </Command>
</LinxMicrovix>
"""


def montar_xml(cnpj, data_ini, data_fim):
    """Monta o XML de requisição para a API Microvix."""
    return XML_TEMPLATE.format(
        usuario=USUARIO,
        senha=SENHA,
        chave=CHAVE,
        cnpj=cnpj,
        data_ini=data_ini,
        data_fim=data_fim,
    )


def parse_xml(xml_texto):
    """Converte o XML de resposta em lista de dicionários."""
    root = ET.fromstring(xml_texto)
    if root.findtext(".//ResponseSuccess") != "True":
        raise Exception(root.findtext(".//ResponseError/Message"))
    colunas = [d.text for d in root.findall(".//ResponseData/C/D")]
    return [
        dict(zip(colunas, [d.text for d in r.findall("D")]))
        for r in root.findall(".//ResponseData/R")
    ]


def para_float(v):
    """Converte string para float, tratando formatos brasileiros."""
    if not v:
        return 0.0
    try:
        return float(v)
    except Exception:
        return float(v.replace(".", "").replace(",", "."))


def buscar_vendas_loja(cnpj, data_str):
    """
    Busca o faturamento do dia para uma loja.
    Usa a mesma lógica de cálculo da Tendência v8.1.
    """
    resp = requests.post(
        URL,
        data=montar_xml(cnpj, data_str, data_str).encode("utf-8"),
        headers={"Content-Type": "application/xml"},
        timeout=60,
    )
    resp.raise_for_status()
    linhas = parse_xml(resp.content.decode("utf-8-sig", errors="replace"))

    faturamento_total = 0.0
    for row in linhas:
        # Ignora registros cancelados ou excluídos
        if (row.get("cancelado") or "").strip().upper() == "S":
            continue
        if (row.get("excluido") or "").strip().upper() == "S":
            continue

        operacao       = (row.get("operacao")       or "").strip().upper()
        tipo_transacao = (row.get("tipo_transacao") or "").strip().upper()

        quantidade    = para_float(row.get("quantidade"))
        valor_liquido = para_float(row.get("valor_liquido"))
        valor_troca   = para_float(row.get("valor_troca"))
        acrescimo     = para_float(row.get("acrescimo"))
        frete         = para_float(row.get("frete"))
        despesas      = para_float(row.get("despesas"))

        valor_final = (
            valor_liquido + valor_troca + acrescimo + frete + despesas
        ) * quantidade

        # Saída (venda) → soma; Devolução de saída → subtrai
        if operacao == "S" and tipo_transacao in ["V", "P", "A", "M", ""]:
            faturamento_total += valor_final
        elif operacao == "DS":
            faturamento_total -= valor_final

    return faturamento_total


def buscar_todas_lojas(data_str):
    """Busca o faturamento de todas as lojas e retorna um dicionário {nome: valor}."""
    vendas = {}
    for cnpj, nome in LOJAS_ATIVAS.items():
        try:
            fat = buscar_vendas_loja(cnpj, data_str)
            vendas[nome] = fat
            status = "COLADA" if fat == 0.0 else f"R$ {fat:,.2f}"
            print(f"  ✓ {nome:<25} {status}")
        except Exception as e:
            print(f"  ✗ {nome:<25} ERRO: {e}")
            vendas[nome] = 0.0
    return vendas


# ============================================================
# ETAPA 2 — MONTA DADOS ORDENADOS
# ============================================================

def montar_dados(vendas):
    """
    Organiza os dados na ordem de exibição:
    lojas Norte → subtotal Norte → lojas Sul → subtotal Sul → total geral.
    """
    todas_lojas = ROTA_NORTE + ROTA_SUL
    resultado   = []

    # Rota Norte
    total_norte = 0.0
    for loja in ROTA_NORTE:
        fat = vendas.get(loja, 0.0)
        total_norte += fat
        resultado.append({"loja": loja, "faturamento": fat, "tipo": "loja"})
    resultado.append({"loja": "ROTA NORTE", "faturamento": total_norte, "tipo": "rota"})

    # Rota Sul
    total_sul = 0.0
    for loja in ROTA_SUL:
        fat = vendas.get(loja, 0.0)
        total_sul += fat
        resultado.append({"loja": loja, "faturamento": fat, "tipo": "loja"})
    resultado.append({"loja": "ROTA SUL", "faturamento": total_sul, "tipo": "rota"})

    # Total geral de todas as lojas
    total_geral = sum(vendas.get(l, 0.0) for l in todas_lojas)
    resultado.append({"loja": "TOTAL GERAL", "faturamento": total_geral, "tipo": "total"})

    return resultado


# ============================================================
# ETAPA 3 — GERA EXCEL
# ============================================================

def criar_excel(dados, output_file):
    """Gera o arquivo .xlsx com formatação de cores por tipo de linha."""
    wb = Workbook()
    ws = wb.active
    ws.title = "PARCIAL DIA"

    # Título e linha em branco
    hoje_fmt = datetime.today().strftime("%d/%m/%Y %H:%M")
    ws["A1"] = f"Parcial do Dia — {hoje_fmt}"
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells("A1:B1")
    ws.append([])

    # Cabeçalho da tabela
    ws.append(["Loja", "Faturamento do Dia"])
    header_row = ws.max_row
    for col in range(1, 3):
        cell           = ws.cell(row=header_row, column=col)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill(start_color="2F4F7F", end_color="2F4F7F", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[header_row].height = 22

    borda_fina   = Border(
        left=Side(style="thin"),  right=Side(style="thin"),
        top=Side(style="thin"),   bottom=Side(style="thin"),
    )
    borda_grossa = Border(
        left=Side(style="thin"),   right=Side(style="thin"),
        top=Side(style="medium"),  bottom=Side(style="medium"),
    )

    for item in dados:
        loja = item["loja"]
        fat  = item["faturamento"]
        tipo = item["tipo"]

        # "COLADA" para lojas sem venda; valor numérico para os demais
        valor_celula = "COLADA" if (tipo == "loja" and fat == 0.0) else fat

        ws.append([loja, valor_celula])
        row = ws.max_row

        if tipo == "total":
            cor   = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            borda = borda_grossa
            fonte = Font(bold=True)
        elif tipo == "rota":
            cor   = PatternFill(start_color="B4C8E6", end_color="B4C8E6", fill_type="solid")
            borda = borda_grossa
            fonte = Font(bold=True, italic=True)
        elif fat == 0.0:
            cor   = PatternFill(start_color="FF6464", end_color="FF6464", fill_type="solid")
            borda = borda_fina
            fonte = Font(bold=True, color="FFFFFF")
        else:
            cor   = PatternFill(start_color="EBF5FF", end_color="EBF5FF", fill_type="solid")
            borda = borda_fina
            fonte = Font()

        for col in range(1, 3):
            cell           = ws.cell(row=row, column=col)
            cell.fill      = cor
            cell.border    = borda
            cell.font      = fonte
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Formato monetário apenas para valores numéricos
        val_cell = ws.cell(row=row, column=2)
        if isinstance(valor_celula, float):
            val_cell.number_format = "R$ #,##0.00"

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22

    wb.save(output_file)
    print(f"  → Excel salvo: {output_file}")


# ============================================================
# ETAPA 4 — GERA IMAGEM PNG
# ============================================================

def carregar_fontes():
    """Tenta carregar uma fonte TTF; usa a padrão se não encontrar."""
    caminhos_bold = [
        "/System/Library/Fonts/Supplemental/Arial Bold.ttf",
        "/System/Library/Fonts/Helvetica.ttc",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "C:/Windows/Fonts/arialbd.ttf",
    ]
    caminhos_normal = [
        "/System/Library/Fonts/Supplemental/Arial.ttf",
        "/System/Library/Fonts/Helvetica.ttc",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "C:/Windows/Fonts/arial.ttf",
    ]
    font_bold = font_normal = None
    for path in caminhos_bold:
        if os.path.exists(path):
            try:
                font_bold = ImageFont.truetype(path, 13)
                break
            except Exception:
                continue
    for path in caminhos_normal:
        if os.path.exists(path):
            try:
                font_normal = ImageFont.truetype(path, 12)
                break
            except Exception:
                continue
    if not font_bold:
        font_bold = ImageFont.load_default()
    if not font_normal:
        font_normal = font_bold
    return font_bold, font_normal


def desenhar_celula(draw, texto, x, y, largura, fonte, cor_texto):
    """Centraliza o texto dentro da célula."""
    bbox = draw.textbbox((0, 0), texto, font=fonte)
    tw   = bbox[2] - bbox[0]
    th   = bbox[3] - bbox[1]
    tx   = x + (largura - tw) // 2
    ty   = y + (ROW_HEIGHT - th) // 2
    draw.text((tx, ty), texto, fill=cor_texto, font=fonte)


def gerar_imagem(dados, output_file):
    """Gera a imagem PNG do relatório para enviar no WhatsApp."""
    font_bold, font_normal = carregar_fontes()
    hoje_fmt = datetime.today().strftime("%d/%m/%Y  %H:%M")

    # Calcula tamanho da imagem: título + info + cabeçalho + linhas de dados
    n_rows  = 2 + 1 + len(dados)
    total_w = PAD_X * 2 + sum(COL_WIDTHS)
    total_h = PAD_Y * 2 + n_rows * ROW_HEIGHT + 16

    img  = Image.new("RGB", (total_w, total_h), COR_BRANCO)
    draw = ImageDraw.Draw(img)
    y    = PAD_Y

    # Linha de título
    draw.text((PAD_X, y + 8), f"Parcial do Dia  —  {hoje_fmt}", fill=COR_HEADER, font=font_bold)
    y += ROW_HEIGHT

    # Linha de subtítulo
    draw.text((PAD_X, y + 8), "Lojas sem vendas aparecem como COLADA", fill=(130, 130, 130), font=font_normal)
    y += ROW_HEIGHT

    # Cabeçalho da tabela
    headers = ["Loja", "Faturamento do Dia"]
    x = PAD_X
    for texto, larg in zip(headers, COL_WIDTHS):
        draw.rectangle([x, y, x + larg, y + ROW_HEIGHT], fill=COR_HEADER)
        draw.rectangle([x, y, x + larg, y + ROW_HEIGHT], outline=COR_BORDA)
        desenhar_celula(draw, texto, x, y, larg, font_bold, COR_TEXTO_HEADER)
        x += larg
    y += ROW_HEIGHT

    # Linhas de dados
    for item in dados:
        loja = item["loja"]
        fat  = item["faturamento"]
        tipo = item["tipo"]
        is_colada = (tipo == "loja" and fat == 0.0)

        if tipo == "total":
            bg, cor_texto, fonte = COR_CINZA, COR_PRETO, font_bold
        elif tipo == "rota":
            bg, cor_texto, fonte = COR_ROTA, COR_PRETO, font_bold
        elif is_colada:
            bg, cor_texto, fonte = COR_COLADA, COR_TEXTO_COLADA, font_bold
        else:
            bg, cor_texto, fonte = COR_VALOR, COR_PRETO, font_normal

        # Linha separadora antes de subtotais e total
        if tipo in ("rota", "total"):
            draw.line([PAD_X, y, PAD_X + sum(COL_WIDTHS), y], fill=(100, 100, 100), width=2)

        # Coluna Loja
        x = PAD_X
        draw.rectangle([x, y, x + COL_WIDTHS[0], y + ROW_HEIGHT], fill=bg)
        draw.rectangle([x, y, x + COL_WIDTHS[0], y + ROW_HEIGHT], outline=COR_BORDA)
        desenhar_celula(draw, loja, x, y, COL_WIDTHS[0], fonte, cor_texto)
        x += COL_WIDTHS[0]

        # Coluna Valor
        texto_val = "COLADA" if is_colada else f"R$ {fat:,.2f}"
        draw.rectangle([x, y, x + COL_WIDTHS[1], y + ROW_HEIGHT], fill=bg)
        draw.rectangle([x, y, x + COL_WIDTHS[1], y + ROW_HEIGHT], outline=COR_BORDA)
        desenhar_celula(draw, texto_val, x, y, COL_WIDTHS[1], fonte, cor_texto)

        y += ROW_HEIGHT

    # Borda externa da tabela
    draw.rectangle(
        [PAD_X, PAD_Y + ROW_HEIGHT * 2, PAD_X + sum(COL_WIDTHS), y],
        outline=(80, 80, 80),
        width=2,
    )

    img.save(output_file)
    print(f"  → Imagem salva: {output_file} ({total_w}x{total_h}px)")


# ============================================================
# ETAPA 5 — ENVIA VIA WHATSAPP
# ============================================================

def enviar_whatsapp(image_path, coladas):
    """Envia a imagem do relatório para todos os destinatários via Evolution API."""
    import base64

    if not EVOLUTION_URL or not EVOLUTION_APIKEY or not EVOLUTION_INSTANCE:
        print("  ⚠ WhatsApp não configurado — verifique EVOLUTION_URL, EVOLUTION_KEY e EVOLUTION_INSTANCIA no .env")
        return

    if not WHATSAPP_DESTINATARIOS:
        print("  ⚠ Nenhum destinatário configurado — verifique WHATSAPP_DESTINATARIOS no .env")
        return

    with open(image_path, "rb") as f:
        imagem_b64 = base64.b64encode(f.read()).decode("utf-8")

    hoje_fmt = datetime.today().strftime("%d/%m/%Y %H:%M")
    coladas_txt = (
        "\n🔴 Coladas: " + ", ".join(coladas)
        if coladas
        else "\n✅ Todas as lojas com vendas!"
    )
    caption = f"📊 *Parcial do Dia - Grupo Talk*\n🕐 {hoje_fmt}{coladas_txt}"

    headers = {"apikey": EVOLUTION_APIKEY, "Content-Type": "application/json"}

    for numero in WHATSAPP_DESTINATARIOS:
        payload = {
            "number": numero,
            "mediatype": "image",
            "mimetype": "image/png",
            "caption": caption,
            "media": imagem_b64,
            "fileName": "PARCIAL_DIA.png",
        }
        try:
            resp = requests.post(
                f"{EVOLUTION_URL}/message/sendMedia/{EVOLUTION_INSTANCE}",
                json=payload,
                headers=headers,
                timeout=60,
            )
            resp.raise_for_status()
            print(f"  → WhatsApp enviado para {numero}")
        except Exception as e:
            print(f"  ✗ Erro ao enviar para {numero}: {e}")


# ============================================================
# MAIN
# ============================================================

def main():
    print("=" * 58)
    print("  PARCIAL DO DIA — GRUPO TALK")
    print("=" * 58)

    # Valida se as credenciais estão configuradas
    if not USUARIO or not SENHA or not CHAVE:
        print("\n❌ Credenciais Microvix não configuradas.")
        print("   Verifique MICROVIX_USER, MICROVIX_SENHA e MICROVIX_CHAVE no .env")
        sys.exit(1)

    hoje     = datetime.today()
    data_str = hoje.strftime("%Y-%m-%d")
    print(f"\nData: {hoje.strftime('%d/%m/%Y')}  |  Hora: {hoje.strftime('%H:%M')}")

    print("\n[1/4] Buscando vendas parciais na API Microvix...")
    vendas = buscar_todas_lojas(data_str)

    print("\n[2/4] Gerando Excel...")
    dados = montar_dados(vendas)
    criar_excel(dados, OUTPUT_EXCEL)

    print("\n[3/4] Gerando imagem PNG...")
    gerar_imagem(dados, OUTPUT_IMAGE)

    print("\n[4/4] Enviando via WhatsApp...")
    coladas = [nome for nome, fat in vendas.items() if fat == 0.0]
    enviar_whatsapp(OUTPUT_IMAGE, coladas)

    print("\n" + "=" * 58)
    print("  Arquivos gerados:")
    print(f"    - {OUTPUT_EXCEL}")
    print(f"    - {OUTPUT_IMAGE}")
    if coladas:
        print(f"\n  Lojas COLADAS ({len(coladas)}): {', '.join(coladas)}")
    print("=" * 58)
    print("\n✅ Parcial do Dia concluído com sucesso!")


if __name__ == "__main__":
    main()
